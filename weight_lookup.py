from openpyxl import load_workbook

FILE_PATH = 'data/weights.xlsx'

def load_headers_and_ids():
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    headers = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]  # skip cow_id
    cow_ids = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    return headers, cow_ids, ws

def find_weight(ws, cow_id, timestamp):
    # Find row for cow_id
    cow_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(cow_id):
            cow_row = row
            break

    if cow_row is None:
        return None

    # Find column for timestamp
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(row=1, column=col).value) == str(timestamp):
            return ws.cell(row=cow_row, column=col).value

    return None

def main():
    try:
        headers, cow_ids, ws = load_headers_and_ids()
    except FileNotFoundError:
        print("Arquivo de dados não encontrado.")
        return
    if not headers:
        print("Nenhum timestamp registrado ainda.")
        return
    if not cow_ids:
        print("Nenhum animal registrado ainda.")
        return

    print("\n🐄 IDs disponíveis:")
    for i, cow in enumerate(cow_ids):
        print(f"{i + 1}. {cow}")
    
    cow_id = input("Digite o ID do animal: ").strip()
    if cow_id not in map(str, cow_ids):
        print("ID inválido.")
        return

    print("\n📅 Pesagens disponíveis:")
    for i, t in enumerate(headers):
        print(f"{i + 1}. {t}")

    try:
        timestamp_index = int(input("Escolha o número da data de pesagem: "))
        timestamp = headers[timestamp_index - 1]
    except (IndexError, ValueError):
        print("Escolha inválida.")
        return

    weight = find_weight(ws, cow_id, timestamp)
    if weight is None:
        print(f"Sem pesagem registrada para o animal {cow_id} na data {timestamp}")
    else:
        print(f"✅ Peso do animal {cow_id} em {timestamp}: {weight} kg")

if __name__ == "__main__":
    main()
