import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

FILE_PATH = 'data/weights.xlsx'

def initialize_excel():
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "weights"
        ws.cell(row=1, column=1, value='cow_id')
        wb.save(FILE_PATH)

def log_weight(cow_id: str, weight_kg: float):
    initialize_excel()
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Check if timestamp already exists in header
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if timestamp not in headers:
        ws.cell(row=1, column=ws.max_column + 1, value=timestamp)
        timestamp_col = ws.max_column
    else:
        timestamp_col = headers.index(timestamp) + 1

    # Check if cow_id already exists
    cow_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(cow_id):
            cow_row = row
            break

    if cow_row is None:
        cow_row = ws.max_row + 1
        ws.cell(row=cow_row, column=1, value=cow_id)

    # Write weight
    ws.cell(row=cow_row, column=timestamp_col, value=weight_kg)
    wb.save(FILE_PATH)
    print(f"Logged {weight_kg}kg for cow {cow_id} at {timestamp}")

# Example usage:
if __name__ == "__main__":
    log_weight("002", 555)
