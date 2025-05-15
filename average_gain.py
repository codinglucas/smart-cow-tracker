from openpyxl import load_workbook
from datetime import datetime

FILE_PATH = 'data/weights.xlsx'

def get_all_data():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    headers = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]  # skip cow_id
    cow_ids = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    return ws, headers, cow_ids

def get_weights_for_cow(ws, cow_id):
    timestamps = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]
    cow_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(cow_id):
            cow_row = row
            break

    weights = []
    if cow_row:
        for col, ts in enumerate(timestamps, start=2):
            value = ws.cell(row=cow_row, column=col).value
            if value is not None:
                try:
                    ts_dt = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
                    weights.append((ts_dt, float(value)))
                except:
                    continue
    return sorted(weights)

def calculate_average_gain(weights):
    if len(weights) < 2:
        return None, None, None, None

    first_date, first_weight = weights[0]
    last_date, last_weight = weights[-1]
    days_between = (last_date - first_date).days

    if days_between == 0:
        return first_date, last_date, first_weight, None

    avg_gain = (last_weight - first_weight) / days_between
    return first_date, last_date, first_weight, avg_gain

def main():
    try:
        ws, headers, cow_ids = get_all_data()
    except FileNotFoundError:
        print("⚠️ Arquivo de dados não encontrado.")
        return

    if not headers or not cow_ids:
        print("⚠️ Nenhum dado disponível.")
        return

    print("\n🐄 IDs disponíveis:")
    for i, cow in enumerate(cow_ids):
        print(f"{i + 1}. {cow}")
    
    cow_id = input("Digite o ID do animal: ").strip()
    if cow_id not in map(str, cow_ids):
        print("ID inválido.")
        return

    weights = get_weights_for_cow(ws, cow_id)
    if len(weights) < 2:
        print("⚠️ São necessárias ao menos duas pesagens para calcular o ganho médio diário.")
        return

    first_date, last_date, first_weight, avg_gain = calculate_average_gain(weights)
    last_weight = weights[-1][1]

    print(f"\n📊 Histórico para o animal {cow_id}")
    print(f"📅 Primeira pesagem: {first_date.strftime('%Y-%m-%d')} — {first_weight:.2f} kg")
    print(f"📅 Última pesagem:   {last_date.strftime('%Y-%m-%d')} — {last_weight:.2f} kg")

    if avg_gain is not None:
        print(f"📈 Ganho médio por dia: {avg_gain:.2f} kg/dia")
    else:
        print("⚠️ As datas das pesagens são no mesmo dia. Não é possível calcular o ganho médio.")

if __name__ == "__main__":
    main()
