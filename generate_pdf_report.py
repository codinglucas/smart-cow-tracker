import os
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

FILE_PATH = 'data/weights.xlsx'
OUTPUT_PDF = 'report.pdf'
TEMP_DIR = 'temp_charts'

def load_data():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    headers = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]
    cow_data = {}
    for row in range(2, ws.max_row + 1):
        cow_id = str(ws.cell(row=row, column=1).value)
        values = []
        for col, ts in enumerate(headers, start=2):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                try:
                    ts_dt = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
                    values.append((ts_dt, float(value)))
                except:
                    continue
        if values:
            cow_data[cow_id] = sorted(values)
    return cow_data

def make_chart(cow_id, values):
    if not os.path.exists(TEMP_DIR):
        os.makedirs(TEMP_DIR)
    
    x = [v[0].strftime('%Y-%m-%d') for v in values]
    y = [v[1] for v in values]
    
    plt.figure(figsize=(6, 4))
    plt.bar(x, y, color='green')
    plt.title(f'Peso do Animal {cow_id}')
    plt.xlabel('Data')
    plt.ylabel('Peso (kg)')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    path = os.path.join(TEMP_DIR, f'{cow_id}.png')
    plt.savefig(path)
    plt.close()
    return path

def generate_pdf(cow_data):
    c = canvas.Canvas(OUTPUT_PDF, pagesize=A4)
    width, height = A4
    margin = 40

    for cow_id, values in cow_data.items():
        img_path = make_chart(cow_id, values)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(margin, height - margin, f"Animal ID: {cow_id}")
        
        img = ImageReader(img_path)
        c.drawImage(img, margin, height/2 - 100, width=width - 2*margin, preserveAspectRatio=True, mask='auto')

        c.showPage()

    c.save()

def cleanup():
    import shutil
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)

def main():
    try:
        cow_data = load_data()
    except FileNotFoundError:
        print("⚠️ Arquivo Excel não encontrado.")
        return
    if not cow_data:
        print("⚠️ Nenhum dado disponível.")
        return

    generate_pdf(cow_data)
    cleanup()
    print(f"✅ Relatório gerado com sucesso: {OUTPUT_PDF}")

def generate_report_from_gui():
    try:
        cow_data = load_data()
    except FileNotFoundError:
        print("⚠️ Arquivo Excel não encontrado.")
        return
    if not cow_data:
        print("⚠️ Nenhum dado disponível.")
        return

    generate_pdf(cow_data)
    cleanup()
    print(f"✅ Relatório gerado com sucesso: {OUTPUT_PDF}")
